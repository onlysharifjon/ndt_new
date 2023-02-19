# print(developers_column)

# if developers_column[0] == ndt_users_dict[5172746353]:
#     # print("xodim aniqlandi")
#     developers_column.insert(int(a) + 1, str(datetime.datetime.now(tz=tzInfo).strftime('%X')))
# print("xodim append qilindi"

# print("start for1")


import datetime

import pytz
import openpyxl
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# set the timezone
tzInfo = pytz.timezone('Asia/Tashkent')
dt = datetime.datetime.now(tz=tzInfo)
developers_column = []
SHETS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
         'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF']
path = "yanvar.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

max_col = sheet_obj.max_column

ndt_users_dict = {1207474771: "Yo`ldoshev Bobur",
                  233029021: "Karimov Anvar",
                  10414033: "Tulaboyev Zafar",
                  2111796525: "Sabina Sobirova",
                  328628941: "Quranboyev Jasur",
                  5172746353: "Sharifjon Mo`minov",
                  520754113: 'Shanazarov Abdullo',
                  524697244: 'Habibullayev Axtam',
                  322626456: 'Сматуллаев Ербол',
                  1336680858: 'Maxmudova Durdona',
                  1755017200: 'Nazaraliyev Jahongir'

                  }

x = datetime.datetime.now()

o = 0
a = x.strftime("%d")

o += 1
for jump in range(10):
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row=jump+1, column=i)
        developers_column.append(cell_obj.value)

    for k in range(32):
        ws[f'{SHETS[k]}{jump+1}'] = developers_column[k]
        print(developers_column[k])
    developers_column.clear()



# print("for ishladi")
# for i in range(1, max_col + 1):
#     cell_obj = sheet_obj.cell(row=1, column=i)
#     developers_column.append(cell_obj.value)
#
# for k in range(32):
#     ws[f'{SHETS[k]}{1}'] = developers_column[k]
#     print(developers_column[k])
# developers_column.clear()
# """---------"""
# for i in range(1, max_col + 1):
#     cell_obj = sheet_obj.cell(row=2, column=i)
#     developers_column.append(cell_obj.value)
# for k in range(32):
#     ws[f'{SHETS[k]}{2}'] = developers_column[k]
#     print(developers_column[k])
# developers_column.clear()
# """---------"""
# for i in range(1, max_col + 1):
#     cell_obj = sheet_obj.cell(row=3, column=i)
#     developers_column.append(cell_obj.value)
# for k in range(32):
#     ws[f'{SHETS[k]}{3}'] = developers_column[k]
#     print(developers_column[k])
# developers_column.clear()
# """---------"""
# for i in range(1, max_col + 1):
#     cell_obj = sheet_obj.cell(row=4, column=i)
#     developers_column.append(cell_obj.value)
# for k in range(32):
#     ws[f'{SHETS[k]}{4}'] = developers_column[k]
#     print(developers_column[k])
# developers_column.clear()
# """---------"""
# for i in range(1, max_col + 1):
#     cell_obj = sheet_obj.cell(row=5, column=i)
#     developers_column.append(cell_obj.value)
# for k in range(32):
#     ws[f'{SHETS[k]}{5}'] = developers_column[k]
#     print(developers_column[k])
# developers_column.clear()




