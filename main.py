import logging
from aiogram.dispatcher import FSMContext
from aiogram import Bot, Dispatcher, types, executor
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.types import CallbackQuery
from aiogram.dispatcher.filters.state import State, StatesGroup
from default import startbut, keldi_xd, ketdi_xd, all_info
from datetime import datetime
from inlines import adminka, ndt_usertable
import datetime
import pytz
import openpyxl
from openpyxl import Workbook
edn = 22
wb = Workbook()
ws = wb.active
# corporation name
corporation = '🏢Next Developers Team'
# set the timezone
tzInfo = pytz.timezone('Asia/Tashkent')
dt = datetime.datetime.now(tz=tzInfo)
developers_column = []
SHETS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
         'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG']
path = "monitoring.xlsx"
# admin ndt users
ADMINSS = [5172746353, 328628941, 233029021]


# clean
class MyStates(StatesGroup):
    next_step = State()
    ketdi_steep = State()
    adminka = State()
    capt = State()
    editer = State()
    editer_date = State()
    time_edit = State()


userr = []
ndt_users_dict = {1207474771: "Yo‘ldoshev Bobur",
                  233029021: "Karimov Anvar",
                  10414033: "Tulaboyev Zafar",
                  2111796525: "Sobirova Sabina",
                  328628941: "Quranboyev Jasur",
                  5172746353: "Sharifjon Mo‘minov",
                  520754113: 'Shanazarov Abdullo',
                  524697244: 'Habibullayev Axtam',
                  322626456: 'Smatullayev Erbol',
                  1336680858: 'Maxmudova Durdona',
                  1755017200: 'Nazaraliyev Jahongir',
                  }
workers_count = 11

API_TOKEN = '5428656747:AAF6y1vkrQ_nFoTBUlJmmRyMELd5RRwt8UA'

XODIMLAR = [5172746353, 328628941, 1207474771, 233029021, 10414033, 2111796525, 520754113,
            524697244, 322626456, 1336680858, 1755017200]
logging.basicConfig(level=logging.INFO)
bot = Bot(token=API_TOKEN, parse_mode='HTML')
dp = Dispatcher(bot, storage=MemoryStorage())
x = datetime.datetime.now()
a = ['Habibullayev Axtam',
     "Quranboyev Jasur",
     "Sabina Sobirovna",
     'Maxmudova Durdona',
     "Shanazarov Abdullo,",
     "Sharifjon Mo‘minov",
     "Smatullayev Erbol",
     "Tulaboyev Zafar",
     "Yo‘ldoshev Bobur",
     ]


@dp.message_handler(commands='edit')
async def edit_data(message: types.Message):
    await message.answer("Foydalanuvchi user id raqamini yuboring")
    await MyStates.editer.set()


@dp.message_handler(state=MyStates.editer, content_types=types.ContentTypes.TEXT)
async def ups(message: types.Message, state: FSMContext):
    usr = message.text
    if type(int(usr)) == type(5):
        await message.answer(f"<b>Xodim</b>: {ndt_users_dict[int(usr)]}")
        await message.answer("Qaysi Sana ma`lumotini o`zgartirmoqchisiz?")
        await MyStates.editer_date.set()
    else:
        await message.answer("Iltimos To`g`ri formatda ma`lumot kiriting!")

    @dp.message_handler(state=MyStates.editer_date, content_types=types.ContentTypes.TEXT)
    async def upg(message: types.Message, state: FSMContext):
        edit_my_date = message.text
        await message.answer("Vaqtini kiriting!")
        await MyStates.time_edit.set()

        @dp.message_handler(state=MyStates.time_edit, content_types=types.ContentTypes.TEXT)
        async def time_edit(message: types.Message, state: FSMContext):
            op = message.text
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            max_col = sheet_obj.max_column
            x = datetime.datetime.now()

            o = 0
            a = x.strftime("%d")
            for jump in range(workers_count):
                for i in range(1, max_col + 1):
                    cell_obj = sheet_obj.cell(row=jump + 1, column=i)
                    developers_column.append(cell_obj.value)
                # print(developers_column)
                if developers_column[0] == ndt_users_dict[int(usr)]:
                    developers_column[int(edit_my_date) + 1] = op
                    await message.answer("<b>Saqlandi</b>")
                for k in range(33):
                    ws[f'{SHETS[k]}{jump + 1}'] = developers_column[k]
                    # print(developers_column[k])
                developers_column.clear()
            for saver in range(33):
                ws.column_dimensions[f'{SHETS[saver]}'].width = 28
            ws.column_dimensions[f'A'].width = 23
            wb.save("monitoring.xlsx")
            await state.finish()


@dp.message_handler(commands='clearmonth')
async def clear_data(message: types.Message):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    x = datetime.datetime.now()
    for saver in range(33):
        ws.column_dimensions[f'{SHETS[saver]}'].width = 28
    ws.column_dimensions[f'A'].width = 23
    a = x.strftime("%d")
    for jump in range(workers_count):
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=jump + 1, column=i)
            developers_column.append(cell_obj.value)
        # print(developers_column)
        print(developers_column)
        for k in range(33):
            if jump == 1:
                ws[f'{SHETS[k]}{jump + 1}'] = developers_column[k]
            elif k <= 2 and jump != 1:
                ws[f'{SHETS[k]}{jump + 1}'] = developers_column[k]
            else:
                ws[f'{SHETS[k]}{jump + 1}'] = "."

            # print(developers_column[k])
        developers_column.clear()
    for d in range(31):
        ws[f'{SHETS[d + 2]}{1}'] = d + 1

    wb.save("monitoring.xlsx")
    await message.answer("Ma`lumotlar tozalandi")


@dp.message_handler(commands=['ndtusers'])
async def send_welcome11(message: types.Message):
    text = f'🧑‍💻 XODIMLAR 🧑‍💻\n\n🧑‍💻{a[0]}\n\n🧑‍💻{a[1]}\n\n👩🏼‍💻{a[2]}\n\n👩🏼‍💻{a[3]}\n\n🧑‍💻{a[4]}\n\n🧑‍💻{a[5]}\n\n🧑‍💻{a[6]}\n\n🧑‍💻{a[7]}\n\n🧑‍💻{a[8]}'
    await message.answer(text)
    a.clear()


"""-----------------------------------------------------------------------adminka-----------------------------------------------------------------"""


@dp.message_handler(commands=['admin'])
async def send_welcome41(message: types.Message):
    if message.from_user.id in ADMINSS:
        await message.delete()
        await message.answer("Menu Admin", reply_markup=adminka)
    else:
        await message.answer("Sizda admin xuquqi mavjud emas")

    @dp.callback_query_handler(text=["month"])
    async def answer1(call: CallbackQuery):
        await call.message.delete()
        ex = open('monitoring.xlsx', 'rb')
        await call.bot.send_document(message.from_user.id, document=ex, caption=corporation)

    @dp.callback_query_handler(text=["week"])
    async def answer2(call: CallbackQuery):
        await call.message.delete()
        await call.message.answer("<b>🧑‍💻Xodimlar ro'yxati</b>", reply_markup=ndt_usertable)
        await message.answer("⬇️Barcha xodimlar ma'lumotlarini ko‘rish⬇️", reply_markup=all_info)


@dp.message_handler(text="📃Barcha xodimlar monitoringi (7 kun)")
async def infos_all(message: types.Message):
    if message.from_user.id in ADMINSS:
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column
        for saver in range(33):
            ws.column_dimensions[f'{SHETS[saver]}'].width = 28
        ws.column_dimensions[f'A'].width = 23
        a = x.strftime("%d")
        for jump in range(8):
            for i in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row=jump + 2, column=i)
                developers_column.append(cell_obj.value)
            # print(developers_column)
            if int(a) >= 7:
                monit_list = []
                for o in range(7):
                    # print(int(a) - 6 + o)
                    if developers_column[int(a) - 6 + o] != '.':
                        monit_list.append(developers_column[int(a) - 6 + o])
                    else:
                        monit_list.append("❌")
                await message.answer(f"""
    <b>Xodim: {developers_column[0]} </b>
    
    <b>{int(a) - 6}:</b> {monit_list[0]}
    
    <b>{int(a) - 5}:</b>  {monit_list[1]}
    
    <b>{int(a) - 4}:</b> {monit_list[2]}
    
    <b>{int(a) - 3}:</b> {monit_list[3]}
    
    <b>{int(a) - 2}:</b> {monit_list[4]}
    
    <b>{int(a) - 1}:</b> {monit_list[5]}
    
    <b>{int(a)}:</b> {monit_list[6]}
    """)
                monit_list.clear()
                developers_column.clear()
            else:
                await message.answer("⏳7 kun ma`lumoti mavjud emas\n\n🕥Bugun Sana")
            # print(f"uz {text}")

    else:
        await message.answer("Sizda admin xuquqi mavjud emas!")


@dp.callback_query_handler(text=["sharif"])
async def answer2(call: CallbackQuery):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    x = datetime.datetime.now()

    o = 0
    a = x.strftime("%d")
    text = ""
    for jump in range(workers_count):
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=jump + 1, column=i)
            developers_column.append(cell_obj.value)
        # print("salom")
        # print(developers_column)
        if developers_column[0] == ndt_users_dict[5172746353]:
            if int(a) >= 7:
                print(True)
                monit_list = []
                for o in range(7):
                    # print(int(a) - 6 + o)
                    if developers_column[int(a) - 6 + o] != '.':
                        monit_list.append(developers_column[int(a) - 6 + o])
                    else:
                        monit_list.append("❌")
                print(True)
                await call.message.answer(f"""
                <b>Xodim: {developers_column[0]} </b>

                <b>{int(a) - 6}:</b> {monit_list[0]}

                <b>{int(a) - 5}:</b>  {monit_list[1]}

                <b>{int(a) - 4}:</b> {monit_list[2]}

                <b>{int(a) - 3}:</b> {monit_list[3]}

                <b>{int(a) - 2}:</b> {monit_list[4]}

                <b>{int(a) - 1}:</b> {monit_list[5]}

                <b>{int(a)}:</b> {monit_list[6]}
                """)
                monit_list.clear()
                print(True)
            else:
                await call.message.answer("⏳7 kun ma`lumoti mavjud emas\n\n🕥Bugun Sana")
        developers_column.clear()


@dp.callback_query_handler(text=["sabina"])
async def answer2(call: CallbackQuery):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    x = datetime.datetime.now()
    o = 0
    a = x.strftime("%d")
    text = ""
    for jump in range(workers_count):
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=jump + 1, column=i)
            developers_column.append(cell_obj.value)
        print(developers_column[0])
        # print("salom")
        # print(developers_column)
        if developers_column[0] == ndt_users_dict[2111796525]:
            print(True)
            if int(a) >= 7:
                monit_list = []
                for o in range(7):
                    # print(int(a) - 6 + o)
                    if developers_column[int(a) - 6 + o] != '.':
                        monit_list.append(developers_column[int(a) - 6 + o])
                    else:
                        monit_list.append("❌")
                await call.message.answer(f"""
                <b>Xodim: {developers_column[0]} </b>

                <b>{int(a) - 6}:</b> {monit_list[0]}

                <b>{int(a) - 5}:</b>  {monit_list[1]}

                <b>{int(a) - 4}:</b> {monit_list[2]}

                <b>{int(a) - 3}:</b> {monit_list[3]}

                <b>{int(a) - 2}:</b> {monit_list[4]}

                <b>{int(a) - 1}:</b> {monit_list[5]}

                <b>{int(a)}:</b> {monit_list[6]}
                """)
                monit_list.clear()

            else:
                await call.message.answer("⏳7 kun ma`lumoti mavjud emas\n\n🕥Bugun Sana")
        developers_column.clear()


@dp.callback_query_handler(text=["jasur"])
async def answer2(call: CallbackQuery):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    x = datetime.datetime.now()

    o = 0
    a = x.strftime("%d")
    text = ""
    for jump in range(workers_count):
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=jump + 1, column=i)
            developers_column.append(cell_obj.value)
        # print("salom")
        # print(developers_column)
        if developers_column[0] == ndt_users_dict[328628941]:
            if int(a) >= 7:
                monit_list = []
                for o in range(7):
                    # print(int(a) - 6 + o)
                    if developers_column[int(a) - 6 + o] != '.':
                        monit_list.append(developers_column[int(a) - 6 + o])
                    else:
                        monit_list.append("❌")
                await call.message.answer(f"""
                <b>Xodim: {developers_column[0]} </b>

                <b>{int(a) - 6}:</b> {monit_list[0]}

                <b>{int(a) - 5}:</b>  {monit_list[1]}

                <b>{int(a) - 4}:</b> {monit_list[2]}

                <b>{int(a) - 3}:</b> {monit_list[3]}

                <b>{int(a) - 2}:</b> {monit_list[4]}

                <b>{int(a) - 1}:</b> {monit_list[5]}

                <b>{int(a)}:</b> {monit_list[6]}
                """)
                monit_list.clear()

            else:
                await call.message.answer("⏳7 kun ma`lumoti mavjud emas\n\n🕥Bugun Sana")
        developers_column.clear()


@dp.callback_query_handler(text=["zafar"])
async def answer2(call: CallbackQuery):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    x = datetime.datetime.now()

    o = 0
    a = x.strftime("%d")
    text = ""
    for jump in range(workers_count):
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=jump + 1, column=i)
            developers_column.append(cell_obj.value)
        # print("salom")
        # print(developers_column)
        if developers_column[0] == ndt_users_dict[10414033]:
            if int(a) >= 7:
                monit_list = []
                for o in range(7):
                    # print(int(a) - 6 + o)
                    if developers_column[int(a) - 6 + o] != '.':
                        monit_list.append(developers_column[int(a) - 6 + o])
                    else:
                        monit_list.append("❌")
                await call.message.answer(f"""
                <b>Xodim: {developers_column[0]} </b>

                <b>{int(a) - 6}:</b> {monit_list[0]}

                <b>{int(a) - 5}:</b>  {monit_list[1]}

                <b>{int(a) - 4}:</b> {monit_list[2]}

                <b>{int(a) - 3}:</b> {monit_list[3]}

                <b>{int(a) - 2}:</b> {monit_list[4]}

                <b>{int(a) - 1}:</b> {monit_list[5]}

                <b>{int(a)}:</b> {monit_list[6]}
                """)
                monit_list.clear()

            else:
                await call.message.answer("⏳7 kun ma`lumoti mavjud emas\n\n🕥Bugun Sana")
        developers_column.clear()


@dp.callback_query_handler(text=["bobur"])
async def answer2(call: CallbackQuery):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    x = datetime.datetime.now()

    o = 0
    a = x.strftime("%d")
    text = ""
    for jump in range(workers_count):
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=jump + 1, column=i)
            developers_column.append(cell_obj.value)
        # print("salom")
        print(developers_column)
        if developers_column[0] == ndt_users_dict[1207474771]:
            if int(a) >= 7:
                monit_list = []
                for o in range(7):
                    # print(int(a) - 6 + o)
                    if developers_column[int(a) - 6 + o] != '.':
                        monit_list.append(developers_column[int(a) - 6 + o])
                    else:
                        monit_list.append("❌")
                await call.message.answer(f"""
                <b>Xodim: {developers_column[0]} </b>

                <b>{int(a) - 6}:</b> {monit_list[0]}

                <b>{int(a) - 5}:</b>  {monit_list[1]}

                <b>{int(a) - 4}:</b> {monit_list[2]}

                <b>{int(a) - 3}:</b> {monit_list[3]}

                <b>{int(a) - 2}:</b> {monit_list[4]}

                <b>{int(a) - 1}:</b> {monit_list[5]}

                <b>{int(a)}:</b> {monit_list[6]}
                """)
                monit_list.clear()
            else:
                await call.message.answer("⏳7 kun ma`lumoti mavjud emas\n\n🕥Bugun Sana")
        developers_column.clear()


@dp.callback_query_handler(text=["abdulla"])
async def answer2(call: CallbackQuery):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    x = datetime.datetime.now()

    o = 0
    a = x.strftime("%d")
    text = ""
    for jump in range(workers_count):
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=jump + 1, column=i)
            developers_column.append(cell_obj.value)
        # print("salom")
        # print(developers_column)
        if developers_column[0] == ndt_users_dict[520754113]:
            if int(a) >= 7:
                monit_list = []
                for o in range(7):
                    # print(int(a) - 6 + o)
                    if developers_column[int(a) - 6 + o] != '.':
                        monit_list.append(developers_column[int(a) - 6 + o])
                    else:
                        monit_list.append("❌")
                await call.message.answer(f"""
                <b>Xodim: {developers_column[0]} </b>

                <b>{int(a) - 6}:</b> {monit_list[0]}

                <b>{int(a) - 5}:</b>  {monit_list[1]}

                <b>{int(a) - 4}:</b> {monit_list[2]}

                <b>{int(a) - 3}:</b> {monit_list[3]}

                <b>{int(a) - 2}:</b> {monit_list[4]}

                <b>{int(a) - 1}:</b> {monit_list[5]}

                <b>{int(a)}:</b> {monit_list[6]}
                """)
                monit_list.clear()

            else:
                await call.message.answer("⏳7 kun ma`lumoti mavjud emas\n\n🕥Bugun Sana")
        developers_column.clear()


@dp.callback_query_handler(text=["durdona"])
async def answer2(call: CallbackQuery):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    x = datetime.datetime.now()

    o = 0
    a = x.strftime("%d")
    text = ""
    for jump in range(workers_count):
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=jump + 1, column=i)
            developers_column.append(cell_obj.value)
        # print("salom")
        # print(developers_column)
        if developers_column[0] == ndt_users_dict[1336680858]:
            if int(a) >= 7:
                monit_list = []
                for o in range(7):
                    # print(int(a) - 6 + o)
                    if developers_column[int(a) - 6 + o] != '.':
                        monit_list.append(developers_column[int(a) - 6 + o])
                    else:
                        monit_list.append("❌")
                await call.message.answer(f"""
                <b>Xodim: {developers_column[0]} </b>

                <b>{int(a) - 6}:</b> {monit_list[0]}

                <b>{int(a) - 5}:</b>  {monit_list[1]}

                <b>{int(a) - 4}:</b> {monit_list[2]}

                <b>{int(a) - 3}:</b> {monit_list[3]}

                <b>{int(a) - 2}:</b> {monit_list[4]}

                <b>{int(a) - 1}:</b> {monit_list[5]}

                <b>{int(a)}:</b> {monit_list[6]}
                """)
                monit_list.clear()

            else:
                await call.message.answer("⏳7 kun ma`lumoti mavjud emas\n\n🕥Bugun Sana")
        developers_column.clear()


@dp.callback_query_handler(text=["jahongir"])
async def answer2(call: CallbackQuery):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    x = datetime.datetime.now()

    o = 0
    a = x.strftime("%d")
    text = ""
    for jump in range(workers_count):
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=jump + 1, column=i)
            developers_column.append(cell_obj.value)
        # print("salom")
        # print(developers_column)
        if developers_column[0] == ndt_users_dict[1755017200]:
            if int(a) >= 7:
                monit_list = []
                for o in range(7):
                    # print(int(a) - 6 + o)
                    if developers_column[int(a) - 6 + o] != '.':
                        monit_list.append(developers_column[int(a) - 6 + o])
                    else:
                        monit_list.append("❌")
                await call.message.answer(f"""
                <b>Xodim: {developers_column[0]} </b>

                <b>{int(a) - 6}:</b> {monit_list[0]}

                <b>{int(a) - 5}:</b>  {monit_list[1]}

                <b>{int(a) - 4}:</b> {monit_list[2]}

                <b>{int(a) - 3}:</b> {monit_list[3]}

                <b>{int(a) - 2}:</b> {monit_list[4]}

                <b>{int(a) - 1}:</b> {monit_list[5]}

                <b>{int(a)}:</b> {monit_list[6]}
                """)
                monit_list.clear()

            else:
                await call.message.answer("⏳7 kun ma`lumoti mavjud emas\n\n🕥Bugun Sana")
        developers_column.clear()


@dp.callback_query_handler(text=["erbol"])
async def answer2(call: CallbackQuery):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    x = datetime.datetime.now()

    o = 0
    a = x.strftime("%d")
    text = ""
    for jump in range(workers_count):
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=jump + 1, column=i)
            developers_column.append(cell_obj.value)
        # print("salom")
        # print(developers_column)
        if developers_column[0] == ndt_users_dict[322626456]:
            if int(a) >= 7:
                monit_list = []
                for o in range(7):
                    # print(int(a) - 6 + o)
                    if developers_column[int(a) - 6 + o] != '.':
                        monit_list.append(developers_column[int(a) - 6 + o])
                    else:
                        monit_list.append("❌")
                await call.message.answer(f"""
                <b>Xodim: {developers_column[0]} </b>

                <b>{int(a) - 6}:</b> {monit_list[0]}

                <b>{int(a) - 5}:</b>  {monit_list[1]}

                <b>{int(a) - 4}:</b> {monit_list[2]}

                <b>{int(a) - 3}:</b> {monit_list[3]}

                <b>{int(a) - 2}:</b> {monit_list[4]}

                <b>{int(a) - 1}:</b> {monit_list[5]}

                <b>{int(a)}:</b> {monit_list[6]}
                """)
                monit_list.clear()

            else:
                await call.message.answer("⏳7 kun ma`lumoti mavjud emas\n\n🕥Bugun Sana")
        developers_column.clear()


@dp.callback_query_handler(text=['axtam'])
async def answer2(call: CallbackQuery):
    print(True)
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    x = datetime.datetime.now()

    o = 0
    a = x.strftime("%d")
    text = ""

    for jump in range(workers_count):
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=jump + 1, column=i)
            developers_column.append(cell_obj.value)
        # print("salom")
        # print(developers_column)
        if developers_column[0] == ndt_users_dict[524697244]:
            print(True)
            if int(a) >= 7:
                monit_list = []
                for o in range(7):
                    # print(int(a) - 6 + o)
                    if developers_column[int(a) - 6 + o] != '.':
                        monit_list.append(developers_column[int(a) - 6 + o])
                    else:
                        monit_list.append("❌")
                await call.message.answer(f"""
                <b>Xodim: {developers_column[0]} </b>

                <b>{int(a) - 6}:</b> {monit_list[0]}

                <b>{int(a) - 5}:</b>  {monit_list[1]}

                <b>{int(a) - 4}:</b> {monit_list[2]}

                <b>{int(a) - 3}:</b> {monit_list[3]}

                <b>{int(a) - 2}:</b> {monit_list[4]}

                <b>{int(a) - 1}:</b> {monit_list[5]}

                <b>{int(a)}:</b> {monit_list[6]}
                """)
                monit_list.clear()

            else:
                await call.message.answer("⏳7 kun ma`lumoti mavjud emas\n\n🕥Bugun Sana")
        developers_column.clear()


"""-----------------------------------------------------------------------adminka-----------------------------------------------------------------"""


@dp.message_handler(commands=['support_admin'])
async def send_photo(message: types.Message):
    await message.answer("Rasmni Jo`nating!")
    await MyStates.adminka.set()


@dp.message_handler(state=MyStates.adminka, content_types=types.ContentTypes.PHOTO)
async def ups(message: types.Message, state: FSMContext):
    photo = message.photo[-1].file_id
    await MyStates.capt.set()
    await message.answer("Rasm qabul qilindi!\n\nRasm uchun text kiriting!")
    await state.finish()
    await MyStates.capt.set()
    print("kaptga zapros bordi")

    @dp.message_handler(state=MyStates.capt, content_types=types.ContentTypes.TEXT)
    async def ool(message: types.Message, state: FSMContext):
        print("capt ga kirdi")
        c = message.text
        await message.answer("Rasm uchun text kiriting!")
        for i in range(len(XODIMLAR)):
            await bot.send_photo(XODIMLAR[i], photo, caption=c)
        await state.finish()


@dp.message_handler(commands=['start'])
async def send_welcome(message: types.Message):
    global user_idtelegram
    xodim_id = message.from_user.id
    await message.answer(
        f"Salom,<b> {message.from_user.full_name}</b>")
    if xodim_id in ADMINSS and XODIMLAR:
        await message.answer("Sizning darajangiz Admin", reply_markup=startbut)
    if xodim_id in XODIMLAR:
        await message.answer(f"<b>{ndt_users_dict[xodim_id]}</b> siz xodimlar bo`limidan foydalana olasiz",
                             reply_markup=startbut)
    if xodim_id not in XODIMLAR:
        await message.answer("Siz bu botdan foydalanishga ruxsat etilmagansiz ❎")


@dp.message_handler(text="Xodim🧑‍💻")
async def sharif(message: types.Message):
    if message.from_user.id in XODIMLAR or ADMINSS:
        await message.answer("🛎 YOQLAMA 🛎", reply_markup=keldi_xd)
    else:
        await message.answer("Botda siz uchun lavozim ajratilmagan !")


@dp.message_handler(state=MyStates.ketdi_steep, content_types=types.ContentTypes.LOCATION)
async def ups(message: types.Message, state: FSMContext):
    print("Ketdiga kirdi")
    await message.forward(233029021, message.message_id, message.chat.id)
    await bot.send_message(233029021,
                           f"🏘<b>Ish vaqti yakunladi !</b>\n💼Xodim: {ndt_users_dict[message.from_user.id]}\n\n🕰Vaqt: {str(datetime.datetime.now(tz=tzInfo).strftime('%X'))}-{str(datetime.datetime.now().strftime('%x'))}")
    # anvar akaga send qilish
    await bot.send_message(328628941,
                           f"🏘<b>Ish vaqti yakunladi !</b>\n💼Xodim: {ndt_users_dict[message.from_user.id]}\n\n🕰Vaqt: {str(datetime.datetime.now(tz=tzInfo).strftime('%X'))}-{str(datetime.datetime.now().strftime('%x'))}")
    await message.forward(328628941, message.message_id, message.chat.id)
    # jasur akaga send qilish
    await message.forward(2111796525, message.message_id, message.chat.id)
    await bot.send_message(2111796525,
                           f"🏘<b>Ish vaqti yakunladi !</b>\n💼Xodim: {ndt_users_dict[message.from_user.id]}\n\n🕰Vaqt: {str(datetime.datetime.now(tz=tzInfo).strftime('%X'))}-{str(datetime.datetime.now().strftime('%x'))}")
    # sabinaga send qilish
    await message.forward(5172746353, message.message_id, message.chat.id)
    await bot.send_message(5172746353,
                           f"🏘<b>Ish vaqti yakunladi !</b>\n💼Xodim: {ndt_users_dict[message.from_user.id]}\n\n🕰Vaqt: {str(datetime.datetime.now(tz=tzInfo).strftime('%X'))}-{str(datetime.datetime.now().strftime('%x'))}")
    # Sharifjonga send qilish

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    x = datetime.datetime.now()
    for saver in range(33):
        ws.column_dimensions[f'{SHETS[saver]}'].width = 28
    ws.column_dimensions[f'A'].width = 23
    a = x.strftime("%d")
    for jump in range(workers_count):
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=jump + 1, column=i)
            developers_column.append(cell_obj.value)
        # print(developers_column)
        if developers_column[0] == ndt_users_dict[message.from_user.id]:
            editer = developers_column[int(a)+1]
            # print(developers_column)
            developers_column[int(a) + 1] = f"""Keldi: {editer} Ketdi: {str(datetime.datetime.now(tz=tzInfo).strftime('%X'))}"""
        for k in range(33):
            ws[f'{SHETS[k]}{jump + 1}'] = developers_column[k]
            # print(developers_column[k])
        developers_column.clear()
    wb.save("monitoring.xlsx")
    await message.answer("Xayr 👋")
    await message.answer("Yangi ish kuniga kech qolmasdan keling 😊")
    await message.answer("Yangi ish kunini Boshlash !", reply_markup=keldi_xd)
    print(f"Ketdi {ndt_users_dict[message.from_user.id]}")
    await state.finish()



@dp.message_handler(text="KELDI 🏢")
async def qoshish(message: types.Message):
    await message.answer("<b>Assalomu Aleykum</b> 😊\nBugungi ishda omad yor bo`lsin.")
    await message.answer("<b>Manzilni Jo`nating!</b>", reply_markup=ketdi_xd)
    await MyStates.next_step.set()


@dp.message_handler(state=MyStates.next_step, content_types=types.ContentTypes.LOCATION)
async def ups(message: types.Message, state: FSMContext):
    print("kirildi keldiga ga")
    await message.answer("📍Manzilingiz Jo`natildi")
    await message.answer("<b>Ish vaqtini yakunlash!💫</b>", reply_markup=ketdi_xd)
    await bot.send_message(233029021,
                           f"🏢<b> ISHGA KELDI</b>\n💼<b>Xodim</b>: {ndt_users_dict[message.from_user.id]}\n\n🕰<b>Vaqt</b>: {str(datetime.datetime.now(tz=tzInfo).strftime('%X'))}-{str(datetime.datetime.now().strftime('%x'))}\n📍Manzil: 👇")
    await message.forward(233029021, message.message_id, message.chat.id)
    # anvar akaga send qilish

    await bot.send_message(2111796525,
                           f"🏢<b> ISHGA KELDI</b>\n💼<b>Xodim</b>: {ndt_users_dict[message.from_user.id]}\n\n🕰<b>Vaqt</b>: {str(datetime.datetime.now(tz=tzInfo).strftime('%X'))}-{str(datetime.datetime.now().strftime('%x'))}\n📍Manzil: 👇")
    await message.forward(2111796525, message.message_id, message.chat.id)
    # sabina opaga send qilish

    await bot.send_message(328628941,
                           f"🏢<b> ISHGA KELDI</b>\n💼<b>Xodim</b>: {ndt_users_dict[message.from_user.id]}\n\n🕰<b>Vaqt</b>: {str(datetime.datetime.now(tz=tzInfo).strftime('%X'))}-{str(datetime.datetime.now().strftime('%x'))}\n📍Manzil: 👇")
    await message.forward(328628941, message.message_id, message.chat.id)
    # jasur aka ga send qilish
    await bot.send_message(5172746353,
                           f"🏢<b> ISHGA KELDI</b>\n💼<b>Xodim</b>: {ndt_users_dict[message.from_user.id]}\n\n🕰<b>Vaqt</b>: {str(datetime.datetime.now(tz=tzInfo).strftime('%X'))}-{str(datetime.datetime.now().strftime('%x'))}\n📍Manzil: 👇")
    await message.forward(5172746353, message.message_id, message.chat.id)

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    x = datetime.datetime.now()

    o = 0
    a = x.strftime("%d")
    for jump in range(workers_count):
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=jump + 1, column=i)
            developers_column.append(cell_obj.value)
        # print(developers_column)
        if developers_column[0] == ndt_users_dict[message.from_user.id]:
            developers_column[int(a) + 1] = str(datetime.datetime.now(tz=tzInfo).strftime('%X'))
            await message.answer("Saqlandi ma`lumot")
        for k in range(len(developers_column)):
            ws[f'{SHETS[k]}{jump + 1}'] = developers_column[k]
            # print(developers_column[k])
        developers_column.clear()
    for saver in range(33):
        ws.column_dimensions[f'{SHETS[saver]}'].width = 28
    ws.column_dimensions[f'A'].width = 23
    wb.save("monitoring.xlsx")
    print(f"Ketdi {ndt_users_dict[message.from_user.id]}")
    await state.finish()



@dp.message_handler(commands='oy_malumotlari')
async def excel_sender(messtage: types.Message):
    ex = open('monitoring.xlsx', 'rb')
    await bot.send_document(messtage.from_user.id, document=ex, caption=corporation)
    # praekt manager send meesage forvard indfo


@dp.message_handler(text="KETDI 🏢")
async def qoshish(message: types.Message):
    await message.answer("Manzilni Tasdiqlang 📍", reply_markup=keldi_xd)
    await MyStates.ketdi_steep.set()


if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)
